import pandas as pd
from openpyxl import load_workbook

# Function to find "System Name" and update the Excel file
def update_excel(excel_file):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel(excel_file)

    # Find the row index where "System Name" is found
    system_name_row = None
    for index, row in df.iterrows():
        if 'System Name' in row.values:
            system_name_row = index
            break

    # If "System Name" is found, prompt the user for input and update Excel file
    if system_name_row is not None:
        user_input = input("Enter text to display beside 'System Name': ")

        # Check if the cell to the right of "System Name" is empty
        if pd.isnull(df.iloc[system_name_row, 1]):  # Assuming "System Name" is found in the first column
            # Update Excel file using openpyxl
            wb = load_workbook(excel_file)
            ws = wb.active

            # Write user input beside "System Name"
            ws.cell(row=system_name_row+1, column=2).value = user_input  # Assuming "System Name" is found in the first column

            # Save the updated Excel file
            wb.save(excel_file)
            print(f"Updated '{excel_file}' with user input: {user_input}")
        else:
            print("Cell beside 'System Name' is not empty. Skipping update.")
    else:
        print("Could not find 'System Name' in the Excel file.")

# Specify the path to your Excel file
excel_file = 'C:/Users/ayoitstp/Desktop/Book1.xlsx'







def main():
    print("hello")
    # Call the function to update the Excel file
    update_excel(excel_file)

if __name__ == "__main__":
    main()