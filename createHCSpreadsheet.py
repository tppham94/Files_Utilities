import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
import pandas as pd 

# Extract devices name from a text file 
def extract_Devices_Names():
    # Specify the file name
    file_name = 'devicesName.txt'

    # Read text file
    with open(file_name, 'r') as file:
        file_contents = file.readlines()
    
    return file_contents

# Creates excel sheet with multiple sheets as needed 
def generate_Sheets_From_List():
    sheet_names = extract_Devices_Names()
    with pd.ExcelWriter('test.xlsx', engine='openpyxl') as writer:
        for sheet_name in sheet_names:
            # Create a simple DataFrame for each sheet (Can be adjusted)
            df = pd.DataFrame({
                'Column1': [1, 2, 3],
                'Column2': ['A', 'B', 'C']
            })
        
            # Write the DataFrame to the specified sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("Excel file created with multiple sheets !")

# Copy template to the newly created spreasheet with the template included 
def copy_Template_Sheet_With_Formatting(source_file, source_sheet_name, target_file, target_sheet_name):
    # Load source workbook and sheet
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet_name]

    # Load target workbook and sheet 
    target_workbook = openpyxl(target_file)
    target_sheet = target_workbook[target_sheet_name]

    # Copy content and formatting 
    for row in source_sheet.iter_rows(min_row=1, max_col=source_sheet.max_column, max_row=source_sheet.max_row):
        for cell in row:
            # Copy value 
            new_cell = target_sheet[cell.coordinate]
            new_cell.value = cell.value

            # Copy fonts
            if cell.font:
                new_cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )
            
            # Copy the fill for cells
            if cell.fill:
                new_cell.fill = PatternFill(
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color,
                    fill_type=cell.fill.fill_type
                )

            # Copy borders 
                new_cell.border = Border(
                    left=Side(style=cell.border.left.style, color=cell.border.left.color),
                    right=Side(style=cell.border.right.style, color=cell.border.right.color),
                    top=Side(style=cell.border.top.style, color=cell.border.top.color),
                    bottom=Side(style=cell.border.bottom.style, color=cell.border.bottom.color)
                )

            # Copy Alignments
            if cell.alignment:
                new_cell.alignment = cell.alignment
            
            # Copy number format
            new_cell.number_format = cell.number_format
    
    # Save target file workbook
    target_workbook.save('test.xlsx')

    # Close the workbooks
    source_workbook.close()
    target_workbook.close()



generate_Sheets_From_List()
# SourceXLSX, SourceSheetName, TargetXLSX, TargetSheetName
# Error here not working 'module' object is not callable 
copy_Template_Sheet_With_Formatting('testTemplate.xlsx', 'template', 'test.xlsx', 'prqsw08a')
# print(extractDevicesNames()